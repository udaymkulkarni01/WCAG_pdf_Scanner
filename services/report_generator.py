"""
Report generator service
Generates HTML and Excel reports from scan results
"""
from pathlib import Path
from typing import List
from datetime import datetime
from jinja2 import Template
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.chart import PieChart, Reference

from utils.logger import setup_logger
from models.scan_result import ScanJob, PDFResult
import config

logger = setup_logger(__name__)


HTML_TEMPLATE = """
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>PDF Compliance Report - {{ job_id }}</title>
    <style>
        * { margin: 0; padding: 0; box-sizing: border-box; }
        body {
            font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            padding: 20px;
            color: #333;
        }
        .container {
            max-width: 1400px;
            margin: 0 auto;
            background: white;
            border-radius: 10px;
            box-shadow: 0 10px 40px rgba(0,0,0,0.2);
            overflow: hidden;
        }
        .header {
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white;
            padding: 30px;
        }
        h1 { font-size: 2em; margin-bottom: 10px; }
        .summary {
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(200px, 1fr));
            gap: 20px;
            padding: 30px;
            background: #f8f9fa;
        }
        .stat-card {
            background: white;
            padding: 20px;
            border-radius: 8px;
            box-shadow: 0 2px 8px rgba(0,0,0,0.1);
            text-align: center;
        }
        .stat-value {
            font-size: 2.5em;
            font-weight: bold;
            margin: 10px 0;
        }
        .stat-label {
            color: #666;
            font-size: 0.9em;
            text-transform: uppercase;
        }
        .compliant { color: #10b981; }
        .non-compliant { color: #ef4444; }
        .error { color: #f59e0b; }
        .results {
            padding: 30px;
        }
        .results h2 {
            margin-bottom: 20px;
            color: #333;
        }
        table {
            width: 100%;
            border-collapse: collapse;
            margin-top: 20px;
            box-shadow: 0 2px 8px rgba(0,0,0,0.1);
            border-radius: 8px;
            overflow: hidden;
        }
        thead {
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white;
        }
        th {
            padding: 15px;
            text-align: left;
            font-weight: 600;
            text-transform: uppercase;
            font-size: 0.85em;
            letter-spacing: 0.5px;
        }
        td {
            padding: 15px;
            border-bottom: 1px solid #e5e7eb;
        }
        tbody tr {
            transition: background-color 0.2s;
        }
        tbody tr:hover {
            background: #f9fafb;
        }
        .status-badge {
            display: inline-block;
            padding: 6px 14px;
            border-radius: 20px;
            font-size: 0.85em;
            font-weight: 600;
            text-transform: uppercase;
            letter-spacing: 0.5px;
        }
        .badge-compliant {
            background: #d1fae5;
            color: #065f46;
        }
        .badge-non-compliant {
            background: #fee2e2;
            color: #991b1b;
        }
        .badge-error {
            background: #fef3c7;
            color: #92400e;
        }
        .view-details-btn {
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white;
            border: none;
            padding: 8px 16px;
            border-radius: 6px;
            cursor: pointer;
            font-size: 0.85em;
            font-weight: 600;
            transition: all 0.3s;
            box-shadow: 0 2px 4px rgba(0,0,0,0.1);
        }
        .view-details-btn:hover {
            transform: translateY(-2px);
            box-shadow: 0 4px 8px rgba(0,0,0,0.2);
        }
        .view-details-btn:active {
            transform: translateY(0);
        }
        .details-row {
            display: none;
            background: #f9fafb;
        }
        .details-row.active {
            display: table-row;
        }
        .details-content {
            padding: 20px;
            animation: slideDown 0.3s ease-out;
        }
        @keyframes slideDown {
            from {
                opacity: 0;
                transform: translateY(-10px);
            }
            to {
                opacity: 1;
                transform: translateY(0);
            }
        }
        .violations-container {
            background: white;
            border-radius: 8px;
            padding: 20px;
            border-left: 4px solid #ef4444;
        }
        .violations-header {
            font-weight: 600;
            margin-bottom: 15px;
            color: #991b1b;
            font-size: 1.1em;
        }
        .violation-item {
            margin: 12px 0;
            padding: 15px;
            background: #fef2f2;
            border-radius: 6px;
            border-left: 3px solid #ef4444;
        }
        .violation-rule {
            font-weight: 700;
            color: #991b1b;
            margin-bottom: 5px;
            font-size: 0.95em;
        }
        .violation-desc {
            color: #555;
            line-height: 1.5;
            margin-bottom: 8px;
        }
        .violation-meta {
            display: flex;
            gap: 20px;
            font-size: 0.85em;
            color: #666;
            margin-top: 8px;
        }
        .violation-meta span {
            background: white;
            padding: 4px 10px;
            border-radius: 4px;
        }
        .no-violations {
            color: #059669;
            font-weight: 600;
            padding: 15px;
            background: #d1fae5;
            border-radius: 6px;
            text-align: center;
        }
        .footer {
            text-align: center;
            padding: 20px;
            background: #f8f9fa;
            color: #666;
            font-size: 0.9em;
        }
    </style>
</head>
<body>
    <div class="container">
        <div class="header">
            <h1>📄 PDF Compliance Report</h1>
            <p>Job ID: {{ job_id }}</p>
            <p>Generated: {{ timestamp }}</p>
        </div>
        
        <div class="summary">
            <div class="stat-card">
                <div class="stat-label">Total PDFs</div>
                <div class="stat-value">{{ total_files }}</div>
            </div>
            <div class="stat-card">
                <div class="stat-label">Compliant</div>
                <div class="stat-value compliant">{{ compliant_count }}</div>
            </div>
            <div class="stat-card">
                <div class="stat-label">Non-Compliant</div>
                <div class="stat-value non-compliant">{{ non_compliant_count }}</div>
            </div>
            <div class="stat-card">
                <div class="stat-label">Success Rate</div>
                <div class="stat-value">{{ "%.1f"|format(success_rate) }}%</div>
            </div>
        </div>
        
        <div class="results">
            <h2>📊 Detailed Results</h2>
            <table>
                <thead>
                    <tr>
                        <th>Filename</th>
                        <th>Status</th>
                        <th>Violations</th>
                        <th>Actions</th>
                    </tr>
                </thead>
                <tbody>
                    {% for result in results %}
                    <tr>
                        <td><strong>{{ result.filename }}</strong></td>
                        <td>
                            <span class="status-badge badge-{{ result.status|lower|replace('_', '-')|replace(' ', '-') }}">
                                {{ result.status }}
                            </span>
                        </td>
                        <td>
                            {% if result.total_violations > 0 %}
                                <span style="color: #ef4444; font-weight: 600;">{{ result.total_violations }} issues</span>
                            {% else %}
                                <span style="color: #10b981; font-weight: 600;">✓ None</span>
                            {% endif %}
                        </td>
                        <td>
                            {% if result.violations %}
                                <button class="view-details-btn" onclick="toggleDetails({{ loop.index0 }})">
                                    👁️ View Details
                                </button>
                            {% else %}
                                <span style="color: #9ca3af;">—</span>
                            {% endif %}
                        </td>
                    </tr>
                    {% if result.violations %}
                    <tr class="details-row" id="details-{{ loop.index0 }}">
                        <td colspan="4">
                            <div class="details-content">
                                <div class="violations-container">
                                    <div class="violations-header">
                                        ⚠️ Accessibility Violations ({{ result.total_violations }})
                                    </div>
                                    {% for v in result.violations %}
                                    <div class="violation-item">
                                        <div class="violation-rule">{{ v.rule_id }}</div>
                                        <div class="violation-desc">{{ v.description }}</div>
                                        <div class="violation-meta">
                                            <span>📋 Specification: {{ v.specification }}</span>
                                            <span>📍 Clause: {{ v.clause }}</span>
                                            <span>❌ Failed Checks: {{ v.failed_checks }}</span>
                                        </div>
                                    </div>
                                    {% endfor %}
                                </div>
                            </div>
                        </td>
                    </tr>
                    {% endif %}
                    {% endfor %}
                </tbody>
            </table>
        </div>
        
        <div class="footer">
            <p>Generated by PDF Compliance Scanner</p>
            <p>Powered by veraPDF</p>
        </div>
    </div>
    
    <script>
        function toggleDetails(index) {
            const detailsRow = document.getElementById('details-' + index);
            const button = event.target;
            
            if (detailsRow.classList.contains('active')) {
                detailsRow.classList.remove('active');
                button.textContent = '👁️ View Details';
            } else {
                detailsRow.classList.add('active');
                button.textContent = '🔼 Hide Details';
            }
        }
        
        // Add keyboard support
        document.addEventListener('keydown', function(e) {
            if (e.key === 'Escape') {
                document.querySelectorAll('.details-row.active').forEach(row => {
                    row.classList.remove('active');
                });
                document.querySelectorAll('.view-details-btn').forEach(btn => {
                    btn.textContent = '👁️ View Details';
                });
            }
        });
    </script>
</body>
</html>
"""


def generate_html_report(job: ScanJob, output_path: str = None) -> str:
    """
    Generate HTML report from scan job.
    
    Args:
        job: ScanJob with scan results
        output_path: Optional output file path
        
    Returns:
        Path to generated HTML file
    """
    logger.info(f"Generating HTML report for job: {job.job_id}")
    
    try:
        # Prepare template data
        template_data = {
            'job_id': job.job_id,
            'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'total_files': job.total_files,
            'compliant_count': job.compliant_count,
            'non_compliant_count': job.non_compliant_count,
            'success_rate': job.success_rate,
            'results': [r.to_dict() for r in job.results]
        }
        
        # Render template
        template = Template(HTML_TEMPLATE)
        html_content = template.render(**template_data)
        
        # Determine output path
        if not output_path:
            output_path = config.REPORTS_FOLDER / f"{job.job_id}.html"
        
        # Write HTML file
        with open(output_path, 'w', encoding='utf-8') as f:
            f.write(html_content)
        
        logger.info(f"✓ HTML report generated: {output_path}")
        return str(output_path)
        
    except Exception as e:
        logger.error(f"Failed to generate HTML report: {e}", exc_info=True)
        raise


def generate_excel_report(job: ScanJob, output_path: str = None) -> str:
    """
    Generate Excel report from scan job.
    
    Args:
        job: ScanJob with scan results
        output_path: Optional output file path
        
    Returns:
        Path to generated Excel file
    """
    logger.info(f"Generating Excel report for job: {job.job_id}")
    
    try:
        # Create workbook
        wb = Workbook()
        
        # Summary sheet
        ws_summary = wb.active
        ws_summary.title = "Summary"
        
        # Header
        ws_summary['A1'] = "PDF Compliance Report"
        ws_summary['A1'].font = Font(size=16, bold=True)
        ws_summary['A2'] = f"Job ID: {job.job_id}"
        ws_summary['A3'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        
        # Statistics
        ws_summary['A5'] = "Metric"
        ws_summary['B5'] = "Value"
        ws_summary['A5'].font = Font(bold=True)
        ws_summary['B5'].font = Font(bold=True)
        
        stats = [
            ("Total PDFs", job.total_files),
            ("Compliant", job.compliant_count),
            ("Non-Compliant", job.non_compliant_count),
            ("Errors", job.error_count),
            ("Success Rate", f"{job.success_rate:.1f}%"),
            ("Duration (sec)", f"{job.duration_seconds:.2f}"),
        ]
        
        for idx, (metric, value) in enumerate(stats, start=6):
            ws_summary[f'A{idx}'] = metric
            ws_summary[f'B{idx}'] = value
        
        # Detailed results sheet
        ws_details = wb.create_sheet("Detailed Results")
        
        # Headers
        headers = ["Filename", "Status", "Profile", "Total Violations", "Failed Checks", "Error"]
        for col, header in enumerate(headers, start=1):
            cell = ws_details.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
        
        # Data rows
        for row_idx, result in enumerate(job.results, start=2):
            ws_details.cell(row=row_idx, column=1, value=result.filename)
            ws_details.cell(row=row_idx, column=2, value=result.status)
            ws_details.cell(row=row_idx, column=3, value=result.profile)
            ws_details.cell(row=row_idx, column=4, value=result.total_violations)
            ws_details.cell(row=row_idx, column=5, value=result.total_failed_checks)
            ws_details.cell(row=row_idx, column=6, value=result.error or "")
            
            # Color code status
            status_cell = ws_details.cell(row=row_idx, column=2)
            if result.compliant:
                status_cell.fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
            elif result.error:
                status_cell.fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
            else:
                status_cell.fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
        
        # Violations sheet
        ws_violations = wb.create_sheet("Violations")
        headers_v = ["Filename", "Rule ID", "Specification", "Description", "Failed Checks"]
        for col, header in enumerate(headers_v, start=1):
            cell = ws_violations.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="EF4444", end_color="EF4444", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
        
        row_idx = 2
        for result in job.results:
            for violation in result.violations:
                ws_violations.cell(row=row_idx, column=1, value=result.filename)
                ws_violations.cell(row=row_idx, column=2, value=violation.rule_id)
                ws_violations.cell(row=row_idx, column=3, value=violation.specification)
                ws_violations.cell(row=row_idx, column=4, value=violation.description)
                ws_violations.cell(row=row_idx, column=5, value=violation.failed_checks)
                row_idx += 1
        
        # Adjust column widths
        for ws in [ws_summary, ws_details, ws_violations]:
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        # Determine output path
        if not output_path:
            output_path = config.REPORTS_FOLDER / f"{job.job_id}.xlsx"
        
        # Save workbook
        wb.save(output_path)
        
        logger.info(f"✓ Excel report generated: {output_path}")
        return str(output_path)
        
    except Exception as e:
        logger.error(f"Failed to generate Excel report: {e}", exc_info=True)
        raise


if __name__ == "__main__":
    print("Testing Report Generator...")
    print("-" * 50)
    print("✓ Report Generator module loaded")
